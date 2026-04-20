# shop/admin_views.py

import os
import tempfile
import requests
from decimal import Decimal
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from shop.models import Product, Category, Brand
import pandas as pd


@staff_member_required
def import_products(request):
    result = None
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        usd_rate = request.POST.get('usd_rate')
        update_existing = request.POST.get('update_existing') == 'yes'
        
        if not excel_file:
            messages.error(request, 'Будь ласка, виберіть файл')
            return redirect('.')
        
        # Отримання курсу USD - ВИПРАВЛЕНО
        if usd_rate and usd_rate.strip():
            try:
                usd_rate = float(usd_rate.replace(',', '.'))
            except:
                usd_rate = None
        else:
            usd_rate = None
        
        if usd_rate is None:
            usd_rate = get_usd_rate()
            if usd_rate is None:
                messages.error(request, 'Не вдалося отримати курс USD. Вкажіть вручну.')
                return redirect('.')
        
        # Зберігаємо файл тимчасово
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        
        try:
            result = process_import(tmp_path, usd_rate, update_existing)
            messages.success(request, f'Імпорт завершено. Створено: {result["created"]}, Оновлено: {result["updated"]}, Помилок: {result["errors"]}')
        except Exception as e:
            messages.error(request, f'Помилка імпорту: {str(e)}')
        finally:
            os.unlink(tmp_path)
        
        return render(request, 'admin/import_products.html', {'result': result})
    
    return render(request, 'admin/import_products.html')


def process_import(file_path, usd_rate, update_existing):
    df = pd.read_excel(file_path, dtype=str)
    
    # Перевірка курсу
    if usd_rate is None:
        usd_rate = 40.0
        print(f"⚠️ Курс USD не вказано, використовуємо {usd_rate}")
    else:
        try:
            usd_rate = float(usd_rate)
            print(f"Використовуємо курс USD: {usd_rate}")
        except:
            usd_rate = 40.0
            print(f"⚠️ Помилка конвертації курсу, використовуємо {usd_rate}")
    
    created = 0
    updated = 0
    errors = 0
    error_list = []
    
    for index, row in df.iterrows():
        try:
            with transaction.atomic():
                # --- Отримання категорії (ВИПРАВЛЕНО) ---
                cat_name = row.get('CategoryName', '')
                
                # Перевірка на пусту назву категорії
                if pd.isna(cat_name) or str(cat_name).strip() == '':
                    cat_name = 'Без категорії'
                else:
                    cat_name = str(cat_name).strip()
                
                # Створюємо slug з назви, якщо назва не пуста
                if cat_name and cat_name != 'Без категорії':
                    slug_value = slugify(cat_name)
                else:
                    slug_value = 'without-category'
                
                # Отримуємо або створюємо категорію
                category, cat_created = Category.objects.get_or_create(
                    name=cat_name,
                    defaults={'slug': slug_value}
                )
                
                if cat_created:
                    print(f"  📁 Створено категорію: {cat_name}")
                
                # --- Бренд ---
                vendor_name = row.get('Vendor', '')
                brand = None
                if not pd.isna(vendor_name) and str(vendor_name).strip() != '':
                    vendor_name = str(vendor_name).strip()
                    brand, _ = Brand.objects.get_or_create(
                        name=vendor_name,
                        defaults={'slug': slugify(vendor_name)}
                    )
                
                # --- Артикул ---
                article = row.get('Article', '')
                if pd.isna(article) or str(article).strip() == '':
                    errors += 1
                    error_list.append(f'Рядок {index+2}: пропущено (немає артикулу)')
                    continue
                article = str(article).strip()
                
                # --- Код ---
                code = row.get('Code', '')
                if pd.isna(code):
                    code = ''
                else:
                    code = str(code).strip()
                
                # --- Ціна USD ---
                price_usd_raw = row.get('PriceUSD', '0')
                if pd.isna(price_usd_raw):
                    price_usd_str = '0'
                else:
                    price_usd_str = str(price_usd_raw).strip().replace(',', '.')
                    import re
                    price_usd_str = re.sub(r'[^\d.-]', '', price_usd_str)
                    if price_usd_str == '' or price_usd_str == '-':
                        price_usd_str = '0'
                
                try:
                    price_usd = Decimal(price_usd_str)
                except:
                    price_usd = Decimal('0')
                
                # Конвертація в UAH
                price_uah = price_usd * Decimal(str(usd_rate))
                price_uah = price_uah.quantize(Decimal('0.01'))
                
                # --- Кількість на складі (quantity) ---
                stock_available = row.get('Stock', '0')
                if pd.isna(stock_available):
                    quantity = 0
                else:
                    stock_str = str(stock_available).strip().lower()
                    try:
                        quantity = int(float(stock_str))
                    except:
                        quantity = 0
                
                # --- Наявність (available) ---
                available = quantity > 0
                
                # --- Гарантія ---
                warranty_val = row.get('Warranty', 0)
                if pd.isna(warranty_val):
                    warranty = 0
                else:
                    try:
                        warranty = int(float(str(warranty_val).replace(',', '.')))
                    except:
                        warranty = 0
                
                # --- Країна ---
                country = row.get('Country', '')
                if pd.isna(country):
                    country = ''
                else:
                    country = str(country).strip()
                
                # --- Назва ---
                title = row.get('Name', '')
                if pd.isna(title):
                    title = 'Без назви'
                else:
                    title = str(title).strip()[:250]
                
                # --- Опис ---
                description = row.get('Description', '')
                if pd.isna(description):
                    description = ''
                else:
                    description = str(description).strip()[:2000]
                
                # --- Додаткові характеристики ---
                attributes = {}
                model = row.get('Model', '')
                if not pd.isna(model) and str(model).strip() != '':
                    attributes['Model'] = str(model).strip()
                
                bonus = row.get('Bonus', '')
                if not pd.isna(bonus) and str(bonus).strip() != '':
                    attributes['Bonus'] = str(bonus).strip()
                
                recommended_price = row.get('RecommendedPrice', '')
                if not pd.isna(recommended_price) and str(recommended_price).strip() != '':
                    attributes['RecommendedPrice'] = str(recommended_price).strip()
                
                # --- Пошук або оновлення ---
                if update_existing:
                    product = Product.objects.filter(article=article).first()
                    if product:
                        product.category = category
                        product.brand = brand
                        product.title = title
                        product.article = article
                        product.code = code
                        product.description = description
                        product.price = price_uah
                        product.quantity = quantity
                        product.available = available
                        product.warranty = warranty
                        product.country = country
                        product.attributes = attributes
                        product.save()
                        updated += 1
                        print(f"  ✅ Оновлено: {article}")
                        continue
                
                # --- Створення нового товару ---
                Product.objects.create(
                    category=category,
                    brand=brand,
                    title=title,
                    article=article,
                    code=code,
                    description=description,
                    price=price_uah,
                    quantity=quantity,
                    available=available,
                    warranty=warranty,
                    country=country,
                    attributes=attributes,
                )
                created += 1
                print(f"  ✅ Створено: {article}")
                
        except Exception as e:
            errors += 1
            error_list.append(f'Рядок {index+2}: {str(e)}')
            print(f"  ❌ ПОМИЛКА: {str(e)}")
    
    print(f"\n=== ПІДСУМОК: Створено: {created}, Оновлено: {updated}, Помилок: {errors} ===")
    
    return {
        'created': created,
        'updated': updated,
        'errors': errors,
        'error_list': error_list,
    }


def get_usd_rate():
    try:
        url = 'https://bank.gov.ua/NBUStatService/v1/statdirectory/exchange?valcode=USD&json'
        response = requests.get(url, timeout=10)
        data = response.json()
        if data and 'rate' in data[0]:
            return float(data[0]['rate'])
    except:
        pass
    return None


def parse_int(value):
    if value is None or str(value).strip() == '' or str(value) == 'nan':
        return 0
    try:
        return int(float(str(value).replace(',', '.')))
    except:
        return 0


def slugify(text):
    from django.utils.text import slugify
    return slugify(text)


@staff_member_required
def export_products(request):
    """Експорт товарів в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товари"
    
    # Заголовки
    headers = [
        'Article', 'Code', 'Name', 'Description', 'PriceUSD', 'CategoryName',
        'Vendor', 'Model', 'Warranty', 'Country', 'Stock', 'Group', 'GroupID',
        'ClassID', 'ClassName', 'CategoryID'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Дані
    products = Product.objects.filter(is_active=True).select_related('category', 'brand')
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.article or '')
        ws.cell(row=row, column=2, value=product.code or '')
        ws.cell(row=row, column=3, value=product.title)
        ws.cell(row=row, column=4, value=product.description or '')
        ws.cell(row=row, column=5, value=float(product.price))
        ws.cell(row=row, column=6, value=product.category.name if product.category else '')
        ws.cell(row=row, column=7, value=product.brand.name if product.brand else '')
        ws.cell(row=row, column=8, value=product.attributes.get('Model', '') if product.attributes else '')
        ws.cell(row=row, column=9, value=product.warranty)
        ws.cell(row=row, column=10, value=product.country or '')
        ws.cell(row=row, column=11, value=1 if product.quantity > 0 else 0)
        ws.cell(row=row, column=12, value=product.attributes.get('Group', '') if product.attributes else '')
        ws.cell(row=row, column=13, value=product.attributes.get('GroupID', '') if product.attributes else '')
        ws.cell(row=row, column=14, value=product.attributes.get('ClassID', '') if product.attributes else '')
        ws.cell(row=row, column=15, value=product.attributes.get('ClassName', '') if product.attributes else '')
        ws.cell(row=row, column=16, value=product.category.external_id.replace('cat_', '') if product.category and product.category.external_id else '')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def export_template(request):
    """Експорт шаблону для імпорту"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шаблон_імпорту"
    
    headers = [
        'CategoryID', 'Code', 'Group', 'Article', 'Vendor', 'Model', 'Name',
        'Description', 'PriceUSD', 'Price_ind', 'CategoryName', 'Bonus',
        'RecommendedPrice', 'DDP', 'Warranty', 'Stock', 'Note', 'DayDelivery',
        'ProductID', 'URL', 'UKTVED', 'GroupID', 'ClassID', 'ClassName',
        'Available', 'Country', 'RetailPrice', 'CostDelivery', 'Exclusive', 'FOP'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 10
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="import_template.xlsx"'
    wb.save(response)
    return response